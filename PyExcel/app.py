import openpyxl

# wb = openpyxl.workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)
# Output ['Sheet1']
sheet = wb["Sheet1"]
print(sheet)
# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)
cell = sheet["a1"]
print(cell.value)
# cell.value = 1
print(cell.row)
print(cell.column)
print(cell.coordinate)
sheet.cell(row=1, column=1)
print(sheet.max_row)
print(sheet.max_column)
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

column = sheet["a"]
print(column)
cells = sheet["a":"c"]
print(cells)
sheet["a1":"c3"]
sheet[1:3]
sheet.append([1, 2, 3])
# sheet.insert_rows()
# sheet.delete_cols()
wb.save("transactions2.xlsx")
